import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.express as px
from collections import Counter
from wordcloud import WordCloud, STOPWORDS
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Crea la cartella images se non esiste
os.makedirs("images", exist_ok=True)

# Contatore globale per i grafici
chart_counter = [0]

def save_fig(fig, width=1200, height=700):
    """Salva il grafico Plotly come PNG nella cartella images/"""
    chart_counter[0] += 1
    filename = f"images/chart_{chart_counter[0]:02d}.png"
    fig.write_image(filename, width=width, height=height, scale=2)
    print(f"  ✅ Salvato: {filename}")

# ============================================================
# CARICAMENTO DATI (generati da Dataset_Importing.py)
# ============================================================
cards = pd.read_csv("cards.csv")

# ============================================================
# CLEANING (da Cleaning_Dataset.py)
# ============================================================
filtered_cards = cards[(cards['type'].isin(['Spell Card', 'Trap Card'])) & (cards['race'].isin(['Continuous', 'Counter','Countinuous','Normal','Equip','Field','Ritual','Quick-Play']))]
filtered_fusion = cards[(cards['type'].isin(['Fusion Monster', 'Link Monster', 'XYZ Monster','Synchro Monster','Synchro Tuner Monster', 'Pendulum Effect Fusion Monster','Synchro Pendulum Effect Monster','Pendulum Effect Fusion Monster','Token','XYZ Pendulum Effect Monster']))] 
filtered_main = cards[(cards['type'].isin(['Effect Monster','Normal Monster','Flip Effect Monster','Union Effect Monster','Pendulum Effect Monster','Tuner Monster','Gemini Monster','Normal Tuner Monster','Spirit Monster','Ritual Effect Monster','Skill Card','Ritual Monster','Toon Monster','Pendulum Normal Monster','Pendulum Tuner Effect Monster','Pendulum Effect Ritual Monster','Pendulum Flip Effect Monster']))] 

banned_cards = cards[(cards['ban_tcg'].isin(['Semi-Limited', 'Limited', 'Banned']))]
banned_cards2 = cards[(cards['ban_ocg'].isin(['Semi-Limited', 'Limited', 'Banned']))]
banned_cards3 = cards[(cards['ban_goat'].isin(['Semi-Limited', 'Limited', 'Banned']))]

cards_na = cards.dropna(subset=['level', 'attribute'])
category_orders = sorted(cards['level'].dropna().unique())

card_set_relation = pd.read_csv("cards_cardsets.csv")
card_set_relation = card_set_relation.drop(columns=["set_id", "set_code"])
card_set_relation = card_set_relation.rename(columns={"card_id": "id"})

# ============================================================
# PREPARAZIONE WRITER EXCEL PER LE TABELLE
# ============================================================
writer = pd.ExcelWriter("tabelle_EDA.xlsx", engine='openpyxl')

# ============================================================
# EDA - GRAFICI
# ============================================================

# 1. Distribuzione dei Livelli delle Carte
fig = px.histogram(cards_na, x='level', color='level', category_orders={'level': category_orders})
fig.update_layout(xaxis_title='Level', yaxis_title='Number of Cards', title_text="Card Level Distribution", legend_title_text='Level', title_x=0.5, width=1200, height=700)
save_fig(fig)

cards_level0 = cards.loc[(cards['level'] == 0)]
print(cards_level0[['name', 'race', 'level','atk', 'def','attribute','type']])

# 2. Confronto ATK vs DEF per livello
fig = px.scatter(cards_na, x='atk', y='def', color='level', symbol='has_effect', hover_name="type")
fig.update_layout(legend=dict(orientation="h"))
fig.update_layout(xaxis_title='Attack Value', yaxis_title='Defense Value', title_text="Attack vs Defense Values by Level", legend_title_text='Has Effect', title_x=0.5, width=1200, height=800)
save_fig(fig)

# 3. Distribuzione Carte Magia e Trappola
fig = px.histogram(filtered_cards, x="race", color="race", facet_col="type")
fig.update_layout(yaxis_title='Number of Cards', title_text="Spell & Trap Card Distribution", legend_title_text='Type', title_x=0.5, width=1500, height=700)
save_fig(fig)

# TABELLA: Upvotes per tipo (Spell/Trap)
somma_per_attributo = filtered_cards.groupby('type')['upvotes'].sum().reset_index()
somma_per_attributo = somma_per_attributo.sort_values(by='upvotes', ascending=False)
somma_per_attributo.to_excel(writer, sheet_name='Upvotes_SpellTrap', index=False)
print(somma_per_attributo)

# 4. Carte più frequenti nell'Extra Deck
custom_palette = {"Fusion Monster": "#b768a2", "Link Monster": "#483d8b", "XYZ Monster": "#1a1a1a", "Synchro Monster": "#c9c9c9", "Synchro Tuner Monster": "#b3b3b3", "Token": "#7a7a7a", "Pendulum Effect Fusion Monster":"#a020f0", "Synchro Pendulum Effect Monster":"#7d26cd", "XYZ Pendulum Effect Monster":"#212121"}
fig = px.histogram(filtered_fusion, x="type", color="type", color_discrete_map=custom_palette)
fig.update_layout(xaxis_title='Type', yaxis_title='Number of Cards', title_text="Most Frequent Cards in the Extra Deck", legend_title_text='Type', title_x=0.5, width=1500, height=700)
fig.update_xaxes(categoryorder='total descending')
save_fig(fig)

# TABELLA: Upvotes per tipo (Extra Deck)
somma_per_attributo1 = filtered_fusion.groupby('type')['upvotes'].sum().reset_index()
somma_per_attributo1 = somma_per_attributo1.sort_values(by='upvotes', ascending=False)
somma_per_attributo1.to_excel(writer, sheet_name='Upvotes_ExtraDeck', index=False)
print(somma_per_attributo1)

# 5. Tipologia di mostro più frequente (Main Deck)
fig = px.histogram(filtered_main, x="type", color="type")
fig.update_layout(yaxis_title='Number of Cards', xaxis_title='Type', title_text="Most Frequent Monster Types (Main Deck)", legend_title_text='Type', title_x=0.5, width=1500, height=700)
fig.update_xaxes(categoryorder='total descending')
save_fig(fig)

# TABELLA: Upvotes per tipo (Main Deck)
somma_per_attributo2 = filtered_main.groupby('type')['upvotes'].sum().reset_index()
somma_per_attributo2 = somma_per_attributo2.sort_values(by='upvotes', ascending=False)
somma_per_attributo2.to_excel(writer, sheet_name='Upvotes_MainDeck', index=False)
print(somma_per_attributo2)

# 6. Attributo più frequente
custom_palette = {"DARK": "#5d3954", "EARTH": "#cd853f", "LIGHT": "#fedf00", "WATER": "#00b2ee", "WIND": "#00ff7f", "FIRE": "#ff0000", "DIVINE":"#ffae42"}
fig = px.histogram(cards_na, x='attribute', color='attribute', color_discrete_map=custom_palette)
fig.update_layout(xaxis_title='Attribute', yaxis_title='Frequency', title_text="Which Attribute is Most and Least Frequent?", legend_title_text='Attribute', title_x=0.5, width=1200, height=700)
fig.update_xaxes(categoryorder='total descending')
save_fig(fig)

# 7. Attributo più preferito (upvotes)
somma_per_attributo3 = cards.groupby('attribute')['upvotes'].sum().reset_index()
somma_per_attributo3 = somma_per_attributo3.sort_values(by='upvotes', ascending=False)

fig = px.histogram(somma_per_attributo3, x='attribute', color='attribute', color_discrete_map=custom_palette, y='upvotes', title="Sum of Upvotes by Attribute")
fig.update_layout(xaxis_title='Attribute', yaxis_title='Frequency', title_text="Which Attribute is Most and Least Preferred?", legend_title_text='Attribute', title_x=0.5, width=1200, height=700)
save_fig(fig)

# 8. Razza più comune
# TABELLA: Upvotes per razza
somma_per_attributo4 = cards_na.groupby('race')['upvotes'].sum().reset_index()
somma_per_attributo4 = somma_per_attributo4.sort_values(by='upvotes', ascending=False)
somma_per_attributo4.to_excel(writer, sheet_name='Upvotes_Razza', index=False)
print(somma_per_attributo4)

fig = px.histogram(cards_na, x='race', color='race')
fig.update_layout(xaxis_title='Race', yaxis_title='Number of Cards', title_text="Which Race is Most and Least Common?", legend_title_text='Race', width=1200, height=800, title_x=0.5)
fig.update_xaxes(tickangle=45, categoryorder='total descending')
save_fig(fig)

# 9. Upvotes vs Views
fig = px.scatter(cards, x='upvotes', y='views', color='type', symbol="has_effect")
fig.update_layout(xaxis_title='Upvotes', yaxis_title='Views', title_text="Do Effect Cards Tend to Be More Rated and Viewed?", legend_title_text='Type', title_x=0.5, width=1200, height=850)
save_fig(fig)

# TABELLA: Top carte per upvotes
filtered_up = cards.loc[(cards['upvotes'] <= 2608) & (cards['upvotes'] >= 423)]
filtered_up = filtered_up[['name', 'upvotes', 'views']].sort_values(by='upvotes', ascending=False)
filtered_up.to_excel(writer, sheet_name='Top_Upvotes', index=False)
print(filtered_up)

# 10. Downvotes vs Views
fig = px.scatter(cards, x='downvotes', y='views', color='type', symbol="has_effect")
fig.update_layout(xaxis_title='Downvotes', yaxis_title='Views', title_text="Are Negatively Rated Effect Cards Also the Most Discussed?", legend_title_text='Type', title_x=0.5, width=1200, height=850)
save_fig(fig)

# TABELLA: Top carte per downvotes
filtered_down = cards.loc[(cards['downvotes'] <= 533) & (cards['downvotes'] >= 109)]
filtered_down = filtered_down[['name', 'downvotes', 'views','ban_tcg','ban_ocg','ban_goat']].sort_values(by='downvotes', ascending=True)
filtered_down.to_excel(writer, sheet_name='Top_Downvotes', index=False)
print(filtered_down)

# TABELLA: Ash Blossom
ash = cards.loc[(cards['name'] == 'Ash Blossom & Joyous Spring')]
ash[['name', 'type', 'desc','attribute','views','upvotes','downvotes','id']].to_excel(writer, sheet_name='Ash_Blossom', index=False)
print(ash[['name', 'type', 'desc','attribute','views','upvotes','downvotes','id']])

# 11. Carte Banate OCG
fig = px.scatter(banned_cards2, x='ban_ocg', y='name', color='type', symbol="has_effect", hover_name="race")
fig.update_layout(xaxis_title='Ban_ocg', yaxis_title='Card Name', title_text="Which Cards Fall into the Forbidden/Limited Category? (OCG)", legend_title_text='Type', title_x=0.5, width=1200, height=850)
fig.update_yaxes(showticklabels=False)
save_fig(fig)

# 12. Carte Banate TCG
fig = px.scatter(banned_cards, x='ban_tcg', y='name', color='type', symbol="has_effect", hover_name="race")
fig.update_layout(xaxis_title='Ban_Tcg', yaxis_title='Card Name', title_text="Which Cards Fall into the Forbidden/Limited Category? (TCG)", legend_title_text='Type', title_x=0.5, width=1200, height=850)
fig.update_yaxes(showticklabels=False)
save_fig(fig)

# Analisi Trap Card
def is_trap_card(description):
    phrases = [
        "This card is also still a Trap", "This card is NOT treated as a Trap",
        "This card is also a Trap Card", "This card is treated as a Trap Card",
        "This card is still treated as a Trap Card", "This card is also still a Trap Card",
        "This card is NOT treated as a Trap Card"
    ]
    return any(phrase in description for phrase in phrases)

trap_cards = cards[cards['desc'].apply(is_trap_card)]
filtered_trap = trap_cards.loc[(trap_cards['race'] == 'Continuous') | (trap_cards['race'] == 'Normal')]
filtered_trap_cols = filtered_trap[['name', 'type', 'desc','race','upvotes']].sort_values(by='name', ascending=True)
filtered_trap_cols.to_excel(writer, sheet_name='Trap_Mostro', index=False)
print(filtered_trap_cols)

# 13. Tipologie di Rarità
merged_data = cards.merge(card_set_relation, on='id')

fig = px.histogram(merged_data, y='set_rarity', color='set_rarity')
fig.update_yaxes(categoryorder='total ascending')
fig.update_layout(title='Types of Card Rarities', xaxis_title='Number of Cards', yaxis_title='Rarity Type', autosize=False, title_x=0.5, width=1500, height=900)
save_fig(fig)

# TABELLA: Rarità speciali
rarita_speciali = merged_data.loc[(merged_data['set_rarity'] == '10000 Secret Rare') | (merged_data['set_rarity'] == 'Ultra Secret Rare') | (merged_data['set_rarity'] == 'c') | (merged_data['set_rarity'] == 'Super') | (merged_data['set_rarity'] == 'Duel Terminal Normal Rare Parallel Rare')]
rarita_speciali[['name', 'type', 'desc','attribute','views','upvotes','downvotes','set_rarity']].to_excel(writer, sheet_name='Rarita_Speciali', index=False)
print(rarita_speciali[['name', 'type', 'desc','attribute','views','upvotes','downvotes','set_rarity']])

# TABELLA: Carte iconiche
carte_iconiche = merged_data.loc[(merged_data['name'] == 'Minerva, the Exalted Lightsworn') | (merged_data['name'] == 'Blue-Eyes White Dragon') | (merged_data['name'] == 'Tyler the Great Warrior') | (merged_data['name'] == 'Dark Magician') | (merged_data['name'] == 'Crush Card Virus') | (merged_data['name'] == 'Cyber-Stein')]
carte_iconiche[['name', 'type', 'desc','attribute','views','upvotes','downvotes','set_rarity','id']].to_excel(writer, sheet_name='Carte_Iconiche', index=False)
print(carte_iconiche[['name', 'type', 'desc','attribute','views','upvotes','downvotes','set_rarity','id']])

# 14. Staple Card Pie Chart
cards['count'] = 1
cards['staple'] = cards['staple'].fillna('No')

staple_counts = cards.groupby('staple')['count'].sum().reset_index().rename(columns={'count': 'Count'})
staple_counts.to_excel(writer, sheet_name='Staple_Conteggio', index=False)
print(staple_counts)

staple_type_counts = cards.groupby(['type', 'staple'])['count'].sum().reset_index()
staple_type_counts = staple_type_counts[staple_type_counts['staple'] == 'Yes'].rename(columns={'count': 'Count'})

custom_palettes = {"Spell Card": "#c9c9c9", "Link Monster": "#b768a2", "XYZ Monster": "#000000", "Effect Monster": "#FF9912", "Tuner Monster": "#b3b3b3", "Trap Card": "#1a1a1a", "Fusion Monster": "#7a7a7a"}
fig = px.pie(staple_type_counts, values='Count', names='type', color_discrete_map=custom_palettes)
fig.update_layout(title="Distribution of Staple Cards by Type", legend_title="Type", title_x=0.5, width=900, height=600)
save_fig(fig)

# TABELLA: Lista Staple Card
staple_list = cards.loc[(cards['staple'] == 'Yes')]
staple_list = staple_list[['name', 'type', 'upvotes', 'views','ban_tcg','ban_ocg','staple']].sort_values(by='upvotes', ascending=True)
staple_list.to_excel(writer, sheet_name='Lista_Staple', index=False)
print(staple_list)

# 15. Treated_as Pie Chart
cards_nas = cards.dropna(subset=['treated_as'])
type_counts = cards_nas['type'].value_counts().reset_index()
type_counts.columns = ['type', 'Count']

fig = px.pie(type_counts, values='Count', names='type')
fig.update_layout(title="Percentage of Treated_as Cards by Type", legend_title="Type", title_x=0.5, width=800, height=800)
save_fig(fig)

cards_nas['treated_as'] = 'Yes'
treated_list = cards_nas[['name', 'type', 'upvotes', 'views','ban_tcg','ban_ocg','treated_as','staple']].sort_values(by='upvotes', ascending=True)
treated_list.to_excel(writer, sheet_name='Lista_TreatedAs', index=False)
print(treated_list)

# 16. Word Cloud
words = pd.read_csv("cards.csv", usecols=["desc"])
text = " ".join(review for review in words.desc)
print("La variabile 'desc' contiene in totale {} parole.".format(len(text)))

all_descriptions = ' '.join(words['desc'].astype(str))
words1 = all_descriptions.split()
word_counts = Counter(words1)
top_words = word_counts.most_common(20)

# TABELLA: Top 20 parole
top_words_df = pd.DataFrame(top_words, columns=['Parola', 'Frequenza'])
top_words_df.to_excel(writer, sheet_name='Top20_Parole', index=False)

for word, count in top_words:
    print(f"La parola {word} è presente {count} volte")

stopwords = set(STOPWORDS)
stopwords.update(["per"])

text = text.lower()
wc = WordCloud(max_words=1000, width=1600, height=800, stopwords=stopwords, contour_width=2.0, contour_color='firebrick', background_color='white')
wc.generate(text)

plt.figure(figsize=[20,10])
plt.imshow(wc, interpolation='bilinear')
plt.axis("off")
plt.savefig("images/wordcloud.png", dpi=150, bbox_inches="tight")
plt.close()
print("  ✅ Salvato: images/wordcloud.png (extra, non nel README)")

# 17. Distribuzione date OCG per serie anime
cards['ocg_date'] = pd.to_datetime(cards['ocg_date'])

date_ranges = [
    ('Yu-Gi-Oh!', '1999-02-04', '2004-12-30'),
    ('Yu-Gi-Oh! GX', '2005-01-20', '2008-03-08'),
    ('Yu-Gi-Oh! 5D', '2008-03-15', '2011-02-26'),
    ('Yu-Gi-Oh! Zexal', '2011-03-19', '2014-03-20'),
    ('Yu-Gi-Oh! Arc-V', '2014-03-21', '2017-03-24'),
    ('Yu-Gi-Oh! VRAINS', '2017-03-25', '2020-04-02'),
    ('Yu-Gi-Oh! Sevens', '2020-04-04', '2023-12-23')
]

sub_datasets = {}
for series, start_date, end_date in date_ranges:
    sub_dataset = cards[(cards['ocg_date'] >= start_date) & (cards['ocg_date'] <= end_date)]
    sub_datasets[series] = sub_dataset

for series, start_date, end_date in date_ranges:
    cards.loc[(cards['ocg_date'] >= start_date) & (cards['ocg_date'] <= end_date), 'serie'] = series

cards = cards.dropna(subset=['serie'])

fig = px.histogram(cards, x='ocg_date', color='serie', nbins=50, title='OCG Release Date Distribution', labels={'ocg_date': 'OCG Date', 'count': 'Number of Cards'})
fig.update_layout(autosize=False, width=1500, height=600, bargap=0.2)
save_fig(fig)

# ============================================================
# FUNZIONE PER ANALISI PER SERIE
# ============================================================
def analisi_serie(nome_serie, sub_data, sheet_prefix):
    print(f"\n{nome_serie} - Numero carte: {len(sub_data)}")
    print(sub_data.isna().sum())

    sub_na = sub_data.dropna(subset=['level', 'attribute'])

    fc = sub_data[(sub_data['type'].isin(['Spell Card', 'Trap Card'])) & (sub_data['race'].isin(['Continuous', 'Counter','Countinuous','Normal','Equip','Field','Ritual','Quick-Play']))]
    ff = sub_data[(sub_data['type'].isin(['Fusion Monster', 'Link Monster', 'XYZ Monster','Synchro Monster','Synchro Tuner Monster', 'Pendulum Effect Fusion Monster','Synchro Pendulum Effect Monster','Pendulum Effect Fusion Monster','Token','XYZ Pendulum Effect Monster']))]
    fm = sub_data[(sub_data['type'].isin(['Effect Monster','Normal Monster','Flip Effect Monster','Union Effect Monster','Pendulum Effect Monster','Tuner Monster','Gemini Monster','Normal Tuner Monster','Spirit Monster','Ritual Effect Monster','Skill Card','Ritual Monster','Toon Monster','Pendulum Normal Monster','Pendulum Tuner Effect Monster','Pendulum Effect Ritual Monster','Pendulum Flip Effect Monster']))]

    # Conteggi Spell/Trap
    cont = fc.groupby(['race', 'type']).size().reset_index(name='conteggio')
    mostri = cont[cont['type'] == 'Spell Card']
    trappole = cont[cont['type'] == 'Trap Card']
    print(f"Magia per razza:\n{mostri}")
    print(f"Trappola per razza:\n{trappole}")
    cont.to_excel(writer, sheet_name=f'{sheet_prefix}_SpellTrap', index=False)

    # Conteggi Extra Deck
    ce = ff['type'].value_counts().reset_index()
    ce.columns = ['Tipo di Carta', 'Conteggio']
    ce = ce.sort_values(by='Conteggio', ascending=False)
    print(ce)
    ce.to_excel(writer, sheet_name=f'{sheet_prefix}_ExtraDeck', index=False)

    # Conteggi Main Deck
    cm = fm['type'].value_counts().reset_index()
    cm.columns = ['Tipo di Carta', 'Conteggio']
    cm = cm.sort_values(by='Conteggio', ascending=False)
    print(cm)
    cm.to_excel(writer, sheet_name=f'{sheet_prefix}_MainDeck', index=False)

    # Grafico Spell/Trap per razza
    if len(fc) > 0:
        spell_data = fc[fc['type'] == 'Spell Card']
        if len(spell_data) > 0:
            fig_spell = px.histogram(spell_data, x='race', color='race')
            fig_spell.update_layout(title=dict(text=f"Number of 'Spell Card' types per race ({nome_serie})", x=0.5), xaxis_title='Race', yaxis_title='Number of Cards', legend_title_text='Race', width=1200, height=700)
            fig_spell.update_xaxes(categoryorder='total descending')
            save_fig(fig_spell)

        trap_data = fc[fc['type'] == 'Trap Card']
        if len(trap_data) > 0:
            fig_trap = px.histogram(trap_data, x='race', color='race')
            fig_trap.update_layout(title=dict(text=f"Number of 'Trap Card' types per race ({nome_serie})", x=0.5), xaxis_title='Race', yaxis_title='Number of Cards', legend_title_text='Race', width=1200, height=700)
            fig_trap.update_xaxes(categoryorder='total descending')
            save_fig(fig_trap)

    # Grafico razza
    if len(sub_na) > 0:
        fig = px.histogram(sub_na, x='race', color='race')
        fig.update_layout(xaxis_title='Race', yaxis_title='Number of Cards', title_text=f"Which Race is Most Frequent? ({nome_serie})", legend_title_text='Race', title_x=0.5, width=1200, height=700)
        fig.update_xaxes(categoryorder='total descending')
        save_fig(fig)

# ============================================================
# ANALISI PER OGNI SERIE
# ============================================================
analisi_serie('Yu-Gi-Oh!', sub_datasets['Yu-Gi-Oh!'], 'YGO')
analisi_serie('Yu-Gi-Oh! GX', sub_datasets['Yu-Gi-Oh! GX'], 'GX')
analisi_serie('Yu-Gi-Oh! 5D', sub_datasets['Yu-Gi-Oh! 5D'], '5D')
analisi_serie('Yu-Gi-Oh! Zexal', sub_datasets['Yu-Gi-Oh! Zexal'], 'Zexal')

analisi_serie('Yu-Gi-Oh! Arc-V', sub_datasets['Yu-Gi-Oh! Arc-V'], 'ArcV')
# Extra chart for Arc-V: Pendulum Scale Distribution
cards_scale = cards.dropna(subset=['scale'])
category_orders = sorted(cards_scale['scale'].unique())
fig = px.histogram(cards_scale, x='scale', color='scale', category_orders={'scale': category_orders})
fig.update_layout(xaxis_title='Scale', yaxis_title='Number of Cards', title_text="Pendulum Monster Scale Distribution", legend_title_text='Scale', title_x=0.5, width=1200, height=700)
save_fig(fig)

analisi_serie('Yu-Gi-Oh! VRAINS', sub_datasets['Yu-Gi-Oh! VRAINS'], 'VRAINS')
# Extra charts for VRAINS: Linkval Distribution + ATK vs Linkval
cards_link = cards.dropna(subset=['linkval','linkmarkers'])
category_orders = sorted(cards_link['linkval'].unique())
fig = px.histogram(cards_link, x='linkval', color='linkval', category_orders={'linkval': category_orders})
fig.update_layout(xaxis_title='Link Value', yaxis_title='Number of Cards', title_text="Link Value Distribution", legend_title_text='Link Value', title_x=0.5, width=1200, height=700)
save_fig(fig)

fig = px.scatter(cards_link, x='atk', y='linkval', color='linkval', symbol='has_effect', hover_name="linkmarkers")
fig.update_layout(legend=dict(orientation="h"))
fig.update_layout(xaxis_title='Attack Value', yaxis_title='Link Value', title_text="Attack Value vs Link Value", legend_title_text='Has Effect', title_x=0.5, width=1200, height=800)
save_fig(fig)

analisi_serie('Yu-Gi-Oh! Sevens', sub_datasets['Yu-Gi-Oh! Sevens'], 'Sevens')

# TABELLA: Skill Cards
skill_cards = cards.loc[(cards['type'] == 'Skill Card')]
skill_cards[['name', 'race', 'level','atk', 'def','formats','type','ocg_date','tcg_date']].to_excel(writer, sheet_name='Skill_Cards', index=False)
print(skill_cards[['name', 'race', 'level','atk', 'def','formats','type','ocg_date','tcg_date']])

# ============================================================
# SALVATAGGIO EXCEL
# ============================================================
writer.close()
print("\n✅ File 'tabelle_EDA.xlsx' salvato con successo!")
print("Fogli contenuti:")
wb = load_workbook("tabelle_EDA.xlsx")
for name in wb.sheetnames:
    print(f"  - {name}")

# ============================================================
# RIEPILOGO
# ============================================================
print(f"\n🎉 Completato! {chart_counter[0]} grafici salvati nella cartella 'images/'")
print("📂 File generati:")
for f in sorted(os.listdir("images")):
    print(f"  📊 images/{f}")